#! python3
# spreadsheetToTxt.py - makes txt files for this example
# Shivan Sivakumaran 22/06/2020

import openpyxl, logging, os

logging.basicConfig(level=logging.DEBUG,format='%(asctime)s - %(levelname)s '
                                               '- %(message)s')
# logging.disable(logging.CRITICAL)

# get the spreadsheet
path = 'Chapter12\\txtToSpreadsheet'
file = os.path.join(path , 'txtToSpreadSheet.xlsx')
wb = openpyxl.load_workbook(file)
sheet = wb.active

# go through spreadsheet
for col in range(1, sheet.max_column+1):
    text_file = open(os.path.join(path, 'from_%s.txt') % (col), 'w')

    for ro in range(1, sheet.max_row+1):
        line = sheet.cell(row=ro, column=col).value
        if line != None:
            text_file.write('%s\n'%line)

    text_file.close()


