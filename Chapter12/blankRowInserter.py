#! python
# blankRowInserter.py <N> <M> <file> - inserts a blank row at location N or size M or a file
# ensure location of file is determined in command line

import openpyxl, sys, os

path = 'Chapter12\\blank'
os.makedirs(path, exist_ok=True)

if len(sys.argv) == 4:
    N = int(sys.argv[1])
    M = int(sys.argv[2])
    file = sys.argv[3]
    filep = os.path.join('Chapter12', file)

print(file)

# open workbook
wb = openpyxl.load_workbook(filep)
sheet = wb.active

# reads up to N lines
if N > sheet.max_row:
    N = sheet.max_row

for row in range(N, N+M):
    for col in range(1, sheet.max_column+1):
        sheet.cell(row=row, column=col).value = ""
        wb.save(os.path.join(path, ('inserted_'+ file)))