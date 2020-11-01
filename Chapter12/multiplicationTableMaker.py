#!python 3
# multiplicationTableMaker.py - enter a number, N in cmd and this results in a table of N size
# by column and row, with numbers multiplied together
# Shivan Sivakumaran

import openpyxl, sys, os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# take arguments from the command prompt
#N = sys.argv[1]
if len(sys.argv) > 1:
    N = int(sys.argv[1])
else:
    N = int(input('Please input size: '))

# build worksheet
wb = openpyxl.Workbook()
sheet = wb.active

# build rows and columns
for i in range(1, N+1):
    fontObj = Font(bold=True)
    sheet['A' + str(i+1)] = i
    sheet['A' + str(i + 1)].font = fontObj


    col = get_column_letter(i + 1)
    sheet[col + '1'] = i
    sheet[col + '1'].font = fontObj

# multiply cells
for i in range(2, sheet.max_row+1):
    for j in range(2, sheet.max_column+1):
        col = get_column_letter(j)
        sheet[col + str(i)] = str(sheet['A' + str(i)].value * sheet[col + '1'].value)

# Save work sheet
os.makedirs('Chapter12\\multiplication', exist_ok=True)
wb.save('Chapter12\\multiplication\\multiplicationTable.xlsx')
print('Sheet Built...')